from __future__ import annotations

from st_aggrid import AgGrid, GridOptionsBuilder,GridUpdateMode, DataReturnMode,JsCode
import streamlit.components.v1 as components
import streamlit as st

import pandas as pd
import numpy as np
import plotly.express as px
import datetime as dt
import textwrap
import html
import hashlib

# To Download Project Checklist 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Iterable, Optional
from io import BytesIO


# # Add to use Java to Auto adjust size of visuals.
on_grid_ready = JsCode("""
function(params) {
    setTimeout(function() {
        params.api.sizeColumnsToFit();
    }, 100);
}
""")

on_grid_size_changed = JsCode("""
function(params) {
    setTimeout(function() {
        params.api.sizeColumnsToFit();
    }, 100);
}
""")


def df_to_excel_bytes(df,
                      sheet_name= "Sheet1",
                      default_max_width= 30,
                      long_columns=[],
                      long_max_width= 80):
    """
    
    
    """
    min_width = 10
    padding = 2
    wrap_vertical_align= "top"
    freeze_header= True,

    long_columns_set = set(long_columns or [])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        safe_sheet = sheet_name[:31]  # Excel sheet name limit
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.sheets[safe_sheet]

        if freeze_header:
            ws.freeze_panes = "A2"

        # Predefine alignments (reuse objects)
        wrap_align = Alignment(wrap_text=True, vertical=wrap_vertical_align)
        no_wrap_align = Alignment(wrap_text=False, vertical=wrap_vertical_align)

        for i, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)

            # Compute max string length in this column (including header)
            ser = df[col].astype(str).fillna("")
            max_len = max(len(str(col)), int(ser.map(len).max()) if len(ser) else 0)

            # Choose cap
            cap = long_max_width if col in long_columns_set else default_max_width

            # Proposed width (with padding)
            proposed = max_len + padding

            # Final width with min + cap
            final_width = max(min_width, min(proposed, cap))
            ws.column_dimensions[col_letter].width = final_width

            # Wrap if we had to cap (meaning content would exceed the allowed width)
            should_wrap = proposed > cap
            if should_wrap:
                # Apply wrap to entire column, incl header
                for cell in ws[col_letter]:
                    cell.alignment = wrap_align
            else:
                # Optional: set vertical alignment consistently
                for cell in ws[col_letter]:
                    cell.alignment = no_wrap_align

    return buffer.getvalue()


def display_reference_grid(
    table_df: pd.DataFrame,
    key: str,
    column_widths: dict[str, int],
    height: int = 350
) -> None:
    """
    Display an AgGrid reference table that refreshes when its data changes.
    """
    grid_df = table_df.copy().reset_index(drop=True)

    missing_columns = [
        column
        for column in column_widths
        if column not in grid_df.columns
    ]

    if missing_columns:
        raise ValueError(
            f"Columns not found in DataFrame: {missing_columns}"
        )

    reference_options = GridOptionsBuilder.from_dataframe(grid_df)

    reference_options.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    for column, width in column_widths.items():
        reference_options.configure_column(
            column,
            width=width,
            minWidth=max(50, int(width * 0.75)),
            maxWidth=int(width * 1.5),
            wrapText=True,
            autoHeight=True
        )

    # Create a hash that changes whenever the DataFrame changes.
    dataframe_hash = hashlib.md5(
        pd.util.hash_pandas_object(
            grid_df,
            index=True
        ).values.tobytes()
    ).hexdigest()

    dynamic_key = f"{key}_{dataframe_hash}"

    AgGrid(
        grid_df,
        gridOptions=reference_options.build(),
        height=height,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
        key=dynamic_key
    )

# ✅ Must be first Streamlit command
st.set_page_config(page_title="Python Function Catalog", layout="wide")

# ✅ Full-width container override
st.markdown(
    """
    <style>
      .block-container {
        max-width: 100% !important;
        padding-left: 1rem;
        padding-right: 1rem;
        padding-top: 0.75rem;
        padding-bottom: 0.5rem;
      }
    </style>
    """,
    unsafe_allow_html=True
)

# -----------------------
# Data sources (raw GitHub)
# -----------------------

@st.cache_data(show_spinner=False)
def load_data():

    
    knowledge_local =       '/Users/derekdewald/Documents/Python/Github_Repo/Streamlit/Data/knowledge_base.xlsx' 
    knowledge_base_xlsx = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/Data/knowledge_base.xlsx"
    
    google_note_csv = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSQF2lNc4WPeTRQ_VzWPkqSZp4RODFkbap8AqmolWp5bKoMaslP2oRVVG21x2POu_JcbF1tGRcBgodu/pub?output=csv'
    google_definition_csv = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vQq1-3cTas8DCWBa2NKYhVFXpl8kLaFDohg0zMfNTAU_Fiw6aIFLWfA5zRem4eSaGPa7UiQvkz05loW/pub?output=csv'
    
    technical_notes = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSnwd-zccEOQbpNWdItUG0qXND5rPVFbowZINjugi15TdWgqiy3A8eMRhbmSMBiRhHt1Qsry3E8tKY8/pub?output=csv'
    function_list = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/Data/python_function_list.csv"
    parameter_list = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/Data/python_function_parameters.csv"

    data_dict = {}
    try:
        data_dict['knowledge_base_df'] = pd.read_excel(knowledge_local)
        print("Local Files Utilized for Knowledge, Consolidated, Process")

    except:
        data_dict['knowledge_base_df'] = pd.read_excel(knowledge_base_xlsx)
        
    data_dict['function_df'] = pd.read_csv(function_list)
    data_dict['google_notes_df'] = pd.read_csv(google_note_csv)
    data_dict['google_definition_df'] = pd.read_csv(google_definition_csv)
    data_dict['technical_notes_df'] = pd.read_csv(technical_notes)
    
    data_dict['parameter_df'] = pd.read_csv(parameter_list)
    
    # Normalize: keep your existing behavior (everything to string)
    for dict_key in data_dict.keys():
        for column in data_dict[dict_key].columns:
            data_dict[dict_key][column] = data_dict[dict_key][column].fillna("").astype(str)

    return data_dict    # Normalize: keep your existing behavior (everything to string)
    for dict_key in data_dict.keys():
        for column in data_dict[dict_key].columns:
            data_dict[dict_key][column] = data_dict[dict_key][column].fillna("").astype(str)

    return data_dict


data_dict = load_data()

# -----------------------
# Navigation
# -----------------------
st.sidebar.title("Navigation")
page = st.sidebar.selectbox(
    "Select Page",
    [ "Home Page", 'Definitions','Notes',"Knowledge Base","Technical Notes",'Processes','Process Checklist','Functions','ML Models','Summarization']
     #"Frequency Summarization",,'Process Checklist',"Function List", "Function Parameters",  'Folder Table of Content', ]
)

if page == "Home Page":
    st.title("Derek's Data Science Knowledge Dasboard")
    st.markdown("""
    <ul>
        <li>The Bedrock of the dashboad is a series of google sheets, .py files maintained on my desktop (and saved in GIT) which represent the approach the processes I follow for work, development, and archival knowledge. The critical pieces are: 
            <ul>
                <li>Definitions</li>
                <li>Notes</li>
                <li>Knowledge Base</li>
                <li>Technical Notes</li>
                <li>Processes</li>
                <li>Processes Checklist</li>
                <li>Functions</li>
            </ul>
        </li>
        <li>Another Main Item</li>
    </ul>
    """, unsafe_allow_html=True)

# # -----------------------------------
# Definitions
# -----------------------------------

elif page == "Definitions":
    st.title("Definitions")

    df_base = data_dict["google_definition_df"].copy()

    # Only convert actual NaN/None to ""
    df_base = df_base.fillna("")

    required = ["Process", "Categorization", "Word", "Definition"]
    missing = [c for c in required if c not in df_base.columns]
    if missing:
        st.error(f"google_definition_df is missing required columns: {missing}")
        st.stop()

    # ----------------------------
    # 1) Slicers
    # ----------------------------
    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        opts1 = ["(All)"] + sorted([x for x in df_base["Process"].astype(str).unique() if str(x).strip()])
        sel1 = st.selectbox("Process", opts1, index=0)

    df1 = df_base if sel1 == "(All)" else df_base[df_base["Process"].astype(str) == str(sel1)]

    with c2:
        opts2 = ["(All)"] + sorted([x for x in df1["Categorization"].astype(str).unique() if str(x).strip()])
        sel2 = st.selectbox("Categorization", opts2, index=0)

    df2 = df1 if sel2 == "(All)" else df1[df1["Categorization"].astype(str) == str(sel2)]

    with c3:
        opts3 = ["(All)"] + sorted([x for x in df2["Word"].astype(str).unique() if str(x).strip()])
        sel3 = st.selectbox("Word", opts3, index=0)

    df_view_full = df2 if sel3 == "(All)" else df2[df2["Word"].astype(str) == str(sel3)]
    st.caption(f"Rows: {len(df_view_full)}")

    # ----------------------------
    # 2) Grid (4 visible cols) + hidden _row_id
    # ----------------------------
    df_view_full = df_view_full.copy().reset_index(drop=False).rename(columns={"index": "_row_id"})


# Build Visual
    visible_cols = ["Process", "Categorization", "Word", "Definition"]
    grid_df = df_view_full[["_row_id"] + visible_cols].copy()

    gb = GridOptionsBuilder.from_dataframe(grid_df)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_selection("single", use_checkbox=False)
    gb.configure_column("_row_id", hide=True)

    gb.configure_column("Process", width=100, minWidth=80, maxWidth=120)
    gb.configure_column("Categorization", width=100, minWidth=80, maxWidth=120)
    gb.configure_column("Word", width=100, minWidth=80, maxWidth=120)

    gb.configure_column(
        "Definition",
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )


    gridOptions = gb.build()
    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed
    gridOptions["domLayout"] = "normal"

    grid_resp = AgGrid(
        grid_df,
        gridOptions=gridOptions,
        height=500,
        fit_columns_on_grid_load=False,
        reload_data=True,
        allow_unsafe_jscode=True,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        )



    selected_rows = grid_resp.get("selected_rows", [])
    if selected_rows is None:
        selected_rows = []
    elif isinstance(selected_rows, pd.DataFrame):
        selected_rows = selected_rows.to_dict("records")

    # ----------------------------
    # 3) Details (HTML-ish rendering)
    # ----------------------------
    st.subheader("Details")

    if len(selected_rows) == 0:
        st.info("Select a row above to view full details.")
    else:
        row_id = selected_rows[0].get("_row_id", None)

        if row_id is None:
            st.warning("Selection did not return _row_id (unexpected).")
        else:
            full_row = df_view_full[df_view_full["_row_id"] == row_id].head(1)

            if full_row.empty:
                st.warning("Could not locate the full record for the selected row.")
            else:
                rec = full_row.iloc[0].fillna("")

                # Show Image (if present)
                if "Image" in rec.index:
                    img_url = str(rec["Image"]).strip()
                    if img_url:
                        st.image(img_url, caption="Image", width=320)

                # Render each field/value like your reference function
                # (Exclude helper + Image since already shown)
                exclude_fields = {"_row_id", "Image"}

                for field, value in rec.items():
                    if field in exclude_fields:
                        continue

                    v = "" if value is None else str(value).strip()

                    # Always show field, even if blank
                    if field.lower() == "link":
                        if v:
                            st.markdown(f"**{field}:** [Open Link]({v})")
                        else:
                            st.markdown(f"**{field}:**")
                    elif field.lower() in {"markdown", "latex"}:
                        st.markdown(f"**{field}:**")
                        if v:
                            try:
                                st.latex(v)
                            except Exception:
                                st.write(v)
                        else:
                            st.write("")
                    else:
                        st.markdown(f"**{field}:**")
                        st.write(v)

# -----------------------------------
# Notes
# -----------------------------------
elif page == 'Notes':
    st.title("Notes")
    df_base = data_dict['google_notes_df'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")
    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_column(c1_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c2_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c3_word, width=150, minWidth=120, maxWidth=170)

    gb.configure_column(
        search_word,
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )

# -----------------------------------
# Knowledge Base
# -----------------------------------
elif page == 'Knowledge Base':
    st.title("Knowledge Base")
    df_base = data_dict['knowledge_base_df'].copy()


    st.subheader("Debug - df_base")
    st.dataframe(df_base)

    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    c4_word = 'Source'
    

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        c4_options = ["(All)"] + sorted([x for x in df3[c4_word].unique() if x.strip()])
        c4_sel = st.selectbox(c4_word, c4_options, index=0)

    df_view = df3 if c4_sel == "(All)" else df3[df3[c4_word] == c4_sel]

    st.caption(f"Rows: {len(df_view)}")

    excel_bytes = df_to_excel_bytes(
        df_view,
        sheet_name="Processes",
        long_columns=["Definition"],
        default_max_width=30,
        long_max_width=80
    )

    st.download_button(
        label="Download filtered Processes as Excel",
        data=excel_bytes,
        file_name="filtered_processes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_column(c1_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c2_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c3_word, width=150, minWidth=120, maxWidth=170)
    gb.configure_column(c4_word, width=150, minWidth=120, maxWidth=170)

    gb.configure_column(
        search_word,
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )

# -----------------------------------
# Technical Notes
# -----------------------------------

elif page == 'Technical Notes':
    st.title("Technical Notes")
    df_base = data_dict['technical_notes_df'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")
    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_column(c1_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c2_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c3_word, width=150, minWidth=120, maxWidth=170)

    gb.configure_column(
        search_word,
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )


# -----------------------------------
# Functions
# -----------------------------------

elif page == 'Functions':
    st.title("Functions")
    df_base = data_dict['function_df'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Function'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df_view = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]
    df_view = df_view[['Folder','Function','Process','Categorization','Definition']]


    st.caption(f"Rows: {len(df_view)}")
    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )

# -----------------------------------
# Processes
# -----------------------------------

elif page == 'ML Models':
    st.title("ML Models")
    df = data_dict['consolidated_df']

    word_list = ['ML Model Taxonomy']
    word_list.extend(df[df['Process'] == 'ML Model Taxonomy']['Word'].dropna().tolist())
    df_base = df[df['Process'].isin(word_list)][['Process','Categorization','Word','Definition']]
    temp = df[df['Process'].isin(df_base['Word'].tolist())][['Process','Categorization','Word','Definition']]
    df_base = pd.concat([df_base,temp]).drop_duplicates(['Word','Process']).fillna("").sort_values(['Process','Categorization','Word'])

    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

############


    # ---------------------------------------------------------
    # Display the three tables side by side
    # ---------------------------------------------------------
    
    # Values in columns are % of screen allocation.
    reference_col1, reference_col2, reference_col3 = st.columns([20,30,50])

    def create_process_summary(source_df,table_=1,selected_word='(All)'):

        # Create a Dataframe for Visualization of Method Objective
        df_table1 = source_df[(source_df['Process']=='Method Objective')&(~source_df['Categorization'].isin(['Process Step','General Definition']))][["Word"]].rename(columns={'Word':"Method Objective"})
        method_objetive_list = df_table1['Method Objective'].tolist()

        # Create a List of ML Model Taxonomy Items for Filtering returned items for Method Graph
        ml_model_tax_df = source_df[(source_df['Process']=='ML Model Taxonomy')][['Word']].rename(columns={'Word':'ML Model Taxonomy'})
        manual_order = {'Learning Paradigm':1,'Method Objective':2,'Method':3,'Method Approach':4}
        ml_model_tax_df['Order'] = ml_model_tax_df['ML Model Taxonomy'].map(manual_order)
        ml_model_tax_df = ml_model_tax_df.sort_values('Order').drop('Order',axis=1).reset_index(drop=True)

        method_approach_df = source_df[source_df['Process']=='Method Approach'][['Word']].rename(columns={'Word':"Method Approach"})

        ml_model_tax_list = ml_model_tax_df['ML Model Taxonomy'].tolist()

        # Create a df of Method Types based on Process = Method. For Other Pertinent Info
        method_df = source_df[(source_df['Process']=='Method')][['Word']].rename(columns={'Word':'Method Types'})

        # Create a Df of Learning Paradigm for Other Pertinent Info.
        lp_df = source_df[source_df['Process']=='Learning Paradigm'][['Word']].rename(columns={'Word':'Learning Paradigm'})

        # Create a Df of Function Type for Other Pertinent Info.
        func_df = source_df[source_df['Process']=='Function'][['Word']].rename(columns={'Word':'Function Types'})

        if table_==1:
            return df_table1

        if table_==2:
            if (selected_word!='(All)')&(selected_word in method_objetive_list):
                return source_df[
                    (source_df['Process'].isin(method_objetive_list))&
                    (source_df['Categorization'].isin(ml_model_tax_list))&
                    (source_df['Process']==selected_word)
                    ][['Word','Process']].rename(columns={'Word':'Method','Process':"Learning Paradigm"})
            else:
                return source_df[(source_df['Process'].isin(method_objetive_list))&(source_df['Categorization'].isin(ml_model_tax_list))][['Word','Process']].rename(columns={'Word':'Method','Process':"Learning Paradigm"})
        
        # Build A Table with ML Taxonomy. Method Types.
        if table_==3:
            # LP - MA - M - MO
            d1 = ml_model_tax_df
            d2 = lp_df.sort_values('Learning Paradigm').reset_index(drop=True)
            d3 = method_approach_df.sort_values('Method Approach').reset_index(drop=True)
            d4 = method_df.sort_values('Method Types').reset_index(drop=True)

            #d5 = func_df.sort_values('Function Types').reset_index(drop=True)
            w = d1.merge(d2,left_index=True,right_index=True,how='outer').merge(d3,left_index=True,right_index=True,how='outer').merge(d4,left_index=True,right_index=True,how='outer').fillna("")

            for col in w.columns:
                if col != "ML Model Taxonomy":
                    w[col] = (
                        w[col]
                        .replace("", pd.NA)
                        .sort_values(ignore_index=True)
                        .fillna("")
                    )
            return w
        
    table_1_df = create_process_summary(df_base)
    table_2_df = create_process_summary(df_base,2,c1_sel)
    table_3_df = create_process_summary(df_base,3)


    #table_3_df = create_process_summary(df_base,3)

    
    with reference_col1:
        st.markdown("#### Method Objective")
    
        display_reference_grid(
            table_1_df,
            column_widths={"Method Objective": 500},
            key="algo_classification_table"
        )

    with reference_col2:
        st.markdown("#### Methods")

        display_reference_grid(
            table_2_df,
            column_widths={'Method':200},
            key="methods_table"
        )

    with reference_col3:
        st.markdown("#### Other Key Information")

        display_reference_grid(
            table_3_df,
            column_widths={'Learning Paradigm':200},
            #column_widths={'ML Model Taxonomy':125,'Method Types':100,"Learning Paradigm":150},
            key="method_type_table"
        )

############




    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_column(c1_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c2_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c3_word, width=150, minWidth=120, maxWidth=170)

    gb.configure_column(
        search_word,
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )


elif page == 'Summarization':
    st.title("Summarization")
    df_base = data_dict['consolidated_df']

    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

############


    # ---------------------------------------------------------
    # Display the three tables side by side
    # ---------------------------------------------------------
    
    # Values in columns are % of screen allocation.
    reference_col1, reference_col2, reference_col3 = st.columns([.25,.25,.5])

    def create_process_summary(source_df,table_=1):

        # Create a Dataframe for Visualization of Algorithm Classification

        words = ['Process','Categorization']

        
        df_table1 = source_df[['Process']].groupby('Process').size().reset_index().rename(columns={0:'Records'})
        df_table2 = source_df[['Categorization']].groupby('Categorization').size().reset_index().rename(columns={0:'Records'})
        df_table3 = source_df[words].groupby(words).size().reset_index().rename(columns={0:'Records'})

        if table_==1:
            return df_table1.sort_values('Records',ascending=False)

        if table_==2:
            return df_table2.sort_values('Records',ascending=False)

        if table_==3:
            return df_table3.sort_values('Records',ascending=False)
        
        
    table_1_df = create_process_summary(df_view)
    table_2_df = create_process_summary(df_view,2)
    table_3_df = create_process_summary(df_view,3)

    with reference_col1:
        st.markdown("#### Learning Paradigm")
    
        display_reference_grid(
            table_1_df,
            column_widths={"Process": 200},
            key="algo_classification_table"
        )

    with reference_col2:
        st.markdown("#### Methods")

        display_reference_grid(
            table_2_df,
            column_widths={'Categorization':200},
            key="methods_table"
        )

    with reference_col3:
        st.markdown("#### Other Key Information")

        display_reference_grid(
            table_3_df,
            column_widths={"Process": 200,'Categorization':200},
            key="method_type_table"
        )

############

    gb = GridOptionsBuilder.from_dataframe(df_view)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        wrapText=True,
        autoHeight=True
    )

    gb.configure_column(c1_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c2_word, width=100, minWidth=80, maxWidth=120)
    gb.configure_column(c3_word, width=150, minWidth=120, maxWidth=170)

    gb.configure_column(
        search_word,
        flex=1,
        minWidth=700,
        wrapText=True,
        autoHeight=True
    )

    gridOptions = gb.build()

    gridOptions["onGridReady"] = on_grid_ready
    gridOptions["onGridSizeChanged"] = on_grid_size_changed

    AgGrid(
        df_view,
        gridOptions=gridOptions,
        height=800,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        reload_data=True,
    )

