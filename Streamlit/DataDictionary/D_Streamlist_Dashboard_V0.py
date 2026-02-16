from __future__ import annotations

from st_aggrid import AgGrid, GridOptionsBuilder,GridUpdateMode
import streamlit.components.v1 as components
import streamlit as st

import pandas as pd
import numpy as np
import plotly.express as px
import datetime as dt
import textwrap
import html


# To Download Project Template 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Iterable, Optional
from io import BytesIO


def df_to_excel_bytes(df,
                      sheet_name= "Sheet1",
                      default_max_width= 30,
                      long_columns=[],
                      long_max_width= 80):
    """
    
    
    """
    min_width = 10
    padding = 2
    wrap_vertical_align= "top"
    freeze_header= True,

    long_columns_set = set(long_columns or [])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        safe_sheet = sheet_name[:31]  # Excel sheet name limit
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.sheets[safe_sheet]

        if freeze_header:
            ws.freeze_panes = "A2"

        # Predefine alignments (reuse objects)
        wrap_align = Alignment(wrap_text=True, vertical=wrap_vertical_align)
        no_wrap_align = Alignment(wrap_text=False, vertical=wrap_vertical_align)

        for i, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)

            # Compute max string length in this column (including header)
            ser = df[col].astype(str).fillna("")
            max_len = max(len(str(col)), int(ser.map(len).max()) if len(ser) else 0)

            # Choose cap
            cap = long_max_width if col in long_columns_set else default_max_width

            # Proposed width (with padding)
            proposed = max_len + padding

            # Final width with min + cap
            final_width = max(min_width, min(proposed, cap))
            ws.column_dimensions[col_letter].width = final_width

            # Wrap if we had to cap (meaning content would exceed the allowed width)
            should_wrap = proposed > cap
            if should_wrap:
                # Apply wrap to entire column, incl header
                for cell in ws[col_letter]:
                    cell.alignment = wrap_align
            else:
                # Optional: set vertical alignment consistently
                for cell in ws[col_letter]:
                    cell.alignment = no_wrap_align

    return buffer.getvalue()

# ✅ Must be first Streamlit command
st.set_page_config(page_title="Python Function Catalog", layout="wide")

# ✅ Full-width container override
st.markdown(
    """
    <style>
      .block-container {
        max-width: 100% !important;
        padding-left: 1rem;
        padding-right: 1rem;
        padding-top: 0.75rem;
        padding-bottom: 0.5rem;
      }
    </style>
    """,
    unsafe_allow_html=True
)

# -----------------------
# Data sources (raw GitHub)
# -----------------------

@st.cache_data(show_spinner=False)
def load_data():
    function_list_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/DataDictionary/python_function_list.csv"
    parameter_list_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/DataDictionary/python_function_parameters.csv"
    folder_toc_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/DataDictionary/folder_listing.csv"
    d_knowledge_base_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Streamlit/DataDictionary/d_knowledge_base.csv"
    notes_def_summary_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Data/d_learning_notes_url_NUMERIC_SUMMARY.csv"
    notes_summary_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Data/google_notes_csv_NUMERIC_SUMMARY.csv"
    def_summary_url = "https://raw.githubusercontent.com/derek-dewald/Python_Tools/main/Data/google_definition_csv_NUMERIC_SUMMARY.csv"
    
    google_note_csv = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSQF2lNc4WPeTRQ_VzWPkqSZp4RODFkbap8AqmolWp5bKoMaslP2oRVVG21x2POu_JcbF1tGRcBgodu/pub?output=csv'
    google_definition_csv = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vQq1-3cTas8DCWBa2NKYhVFXpl8kLaFDohg0zMfNTAU_Fiw6aIFLWfA5zRem4eSaGPa7UiQvkz05loW/pub?output=csv'
    google_word_quote = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vTjXiFjpGgyqWDg9RImj1HR_BeriXs4c5-NSJVwQFn2eRKksitY46oJT0GvVX366LO-m1GM8znXDcBp/pub?gid=1117793378&single=true&output=csv'
    technical_notes = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSnwd-zccEOQbpNWdItUG0qXND5rPVFbowZINjugi15TdWgqiy3A8eMRhbmSMBiRhHt1Qsry3E8tKY8/pub?output=csv'
    
    data_dict = {}

    data_dict['google_notes_df'] = pd.read_csv(google_note_csv)
    data_dict['google_definition_df'] = pd.read_csv(google_definition_csv)
    data_dict['function_list_df'] = pd.read_csv(function_list_url)
    data_dict['parameter_list_df'] = pd.read_csv(parameter_list_url)
    data_dict['folder_toc_df'] = pd.read_csv(folder_toc_url)
    data_dict['d_knowledge_base'] = pd.read_csv(d_knowledge_base_url)
    data_dict['d_knowledge_base'] = data_dict['d_knowledge_base'][['Process','Categorization','Word','Definition']]
    data_dict['notes_def_summary'] = pd.read_csv(notes_def_summary_url)
    data_dict['notes_summary'] = pd.read_csv(notes_summary_url)
    data_dict['def_summary'] = pd.read_csv(def_summary_url)
    data_dict['d_word_quote'] = pd.read_csv(google_word_quote)
    data_dict['technical_notes'] = pd.read_csv(technical_notes)
    

    # ✅ Folder first, then Function
    data_dict['function_list_df1']  = data_dict['function_list_df'][["Folder", "Function", "Purpose"]].copy()
    data_dict['parameter_list_df1'] = data_dict['parameter_list_df'][["Folder", "Function", "Parameters", "Definition"]].copy()

    # Normalize: keep your existing behavior (everything to string)
    for dict_key in data_dict.keys():
        for column in data_dict[dict_key].columns:
            data_dict[dict_key][column] = data_dict[dict_key][column].fillna("").astype(str)

    return data_dict

data_dict = load_data()

# -----------------------
# Navigation
# -----------------------
st.sidebar.title("Navigation")
page = st.sidebar.selectbox(
    "Select Page",
    [ "Knowledge Base",'D Notes', 'D Definitions',"Frequency Summarization",'Technical Notes','Words and Quotes','Project Template',
     "Function List", "Function Parameters",  'Folder Table of Content', 
     ]
)

# -------------------------
# Function List
# -------------------------
if page == "Function List":
    st.title("Function List")
    df_base = data_dict['function_list_df1'].copy()
    c1, c2, c3 = st.columns([1, 1, 2])

    with c1:
        folder_opts = ["(All)"] + sorted([x for x in df_base["Folder"].unique() if x.strip()])
        sel_folder = st.selectbox("Folder", folder_opts, index=0)

    df1 = df_base if sel_folder == "(All)" else df_base[df_base["Folder"] == sel_folder]

    with c2:
        func_opts = ["(All)"] + sorted([x for x in df1["Function"].unique() if x.strip()])
        sel_func = st.selectbox("Function", func_opts, index=0)

    df2 = df1 if sel_func == "(All)" else df1[df1["Function"] == sel_func]

    with c3:
        purpose_search = st.text_input("Purpose search", value="", placeholder="Type to search Purpose...")

    df_view = df2
    if purpose_search.strip():
        s = purpose_search.strip().lower()
        df_view = df_view[df_view["Purpose"].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_column("Folder", width=100)
    gb.configure_column("Function", width=100)
    gb.configure_column("Purpose", flex=1, wrapText=True, autoHeight=True)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)

    AgGrid(df_view, gridOptions=gb.build(), height=800, fit_columns_on_grid_load=True)


# -----------------------------------
# Function Parameters
# -----------------------------------
elif page == "Function Parameters":
    st.title("Function Parameters")
    df_base = data_dict['parameter_list_df1'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    with c1:
        folder_opts = ["(All)"] + sorted([x for x in df_base["Folder"].unique() if x.strip()])
        sel_folder = st.selectbox("Folder", folder_opts, index=0)

    df1 = df_base if sel_folder == "(All)" else df_base[df_base["Folder"] == sel_folder]

    with c2:
        func_opts = ["(All)"] + sorted([x for x in df1["Function"].unique() if x.strip()])
        sel_func = st.selectbox("Function", func_opts, index=0)

    df2 = df1 if sel_func == "(All)" else df1[df1["Function"] == sel_func]

    with c3:
        param_opts = ["(All)"] + sorted([x for x in df2["Parameters"].unique() if x.strip()])
        sel_param = st.selectbox("Parameters", param_opts, index=0)

    df3 = df2 if sel_param == "(All)" else df2[df2["Parameters"] == sel_param]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Definition...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view["Definition"].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_column("Folder", width=180)
    gb.configure_column("Function", width=220)
    gb.configure_column("Parameters", width=320)
    gb.configure_column("Definition", flex=1, wrapText=True, autoHeight=True)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)

    AgGrid(df_view, gridOptions=gb.build(), height=800, fit_columns_on_grid_load=True)

# -----------------------------------
# D Notes
# -----------------------------------
elif page == 'D Notes':
    st.title("D Notes")
    df_base = data_dict['google_notes_df'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_column(c1_word, width=100)
    gb.configure_column(c2_word, width=100)
    gb.configure_column(c3_word, width=150)
    gb.configure_column(search_word, flex=1, wrapText=True, autoHeight=True)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)

    AgGrid(df_view, gridOptions=gb.build(), height=800, fit_columns_on_grid_load=True)


# -----------------------------------
# Folder Table of Content
# -----------------------------------
elif page == "Folder Table of Content":
    st.title("Folder Table of Content")
    df_base = data_dict['folder_toc_df'].copy()
    if "Type" in df_base.columns:
        df_base.drop('Type', inplace=True, axis=1)
    st.write(df_base)

# -----------------------------------
# D Notes Outline
# -----------------------------------
elif page == "Knowledge Base":
    st.title("Knowledge Base")
    df_base = data_dict['d_knowledge_base'].copy()
    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])

    c1_word = 'Process'
    c2_word = 'Categorization'
    c3_word = 'Word'
    search_word = 'Definition'

    with c1:
        c1_options = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        c1_sel = st.selectbox(c1_word, c1_options, index=0)

    df1 = df_base if c1_sel == "(All)" else df_base[df_base[c1_word] == c1_sel]

    with c2:
        c2_options = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        c2_sel = st.selectbox(c2_word, c2_options, index=0)

    df2 = df1 if c2_sel == "(All)" else df1[df1[c2_word] == c2_sel]

    with c3:
        c3_options = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        c3_sel = st.selectbox(c3_word, c3_options, index=0)

    df3 = df2 if c3_sel == "(All)" else df2[df2[c3_word] == c3_sel]

    with c4:
        definition_search = st.text_input("Definition search", value="", placeholder="Type to search Description...")

    df_view = df3
    if definition_search.strip():
        s = definition_search.strip().lower()
        df_view = df_view[df_view[search_word].str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_column(c1_word, width=100)
    gb.configure_column(c2_word, width=100)
    gb.configure_column(c3_word, width=150)
    gb.configure_column(search_word, flex=1, wrapText=True, autoHeight=True)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)

    AgGrid(df_view, gridOptions=gb.build(), height=800, fit_columns_on_grid_load=True)

# -----------------------------------
# Words and Quotes
# -----------------------------------

elif page == "Words and Quotes":
    st.title("Words and Quotes")
    df_base = data_dict["d_word_quote"].copy()
    df_base = df_base[(df_base['Text'].notnull()) & (df_base['Text'] != "")].copy()

    # --- Robust date parsing (handles strings + true date values) ---
    if "Date" in df_base.columns:
        # First try strict known format, then fallback
        dt1 = pd.to_datetime(df_base["Date"], format="%d-%b-%y", errors="coerce")
        dt2 = pd.to_datetime(df_base["Date"], errors="coerce")
        df_base["Date_dt"] = dt1.fillna(dt2)

        # A consistent label for display + selection
        df_base["Date_label"] = df_base["Date_dt"].dt.strftime("%d-%b-%y")

        # Sort newest first
        df_base = df_base.sort_values("Date_dt", ascending=False)

    c1_word = "Date"
    c2_word = "Item"
    c3_word = "Source"
    c4_word = "Chapter"
    c5_word = "Verse(s)"
    search_word = "Text"

    c1, c2, c3, c4, c5, c6 = st.columns([1, 1, 1, 1, 1, 2])

    # --- Date slicer uses Date_label, filters on Date_dt ---
    with c1:
        opts1 = ["(All)"] + [
            x for x in df_base["Date_label"].dropna().unique().tolist()
            if str(x).strip()
        ]
        # keep options sorted by actual datetime (not string)
        opts1_sorted = ["(All)"] + (
            df_base.dropna(subset=["Date_dt"])
                  .drop_duplicates("Date_label")
                  .sort_values("Date_dt", ascending=False)["Date_label"]
                  .tolist()
        )
        sel1 = st.selectbox("Date", opts1_sorted, index=0)

    df1 = df_base if sel1 == "(All)" else df_base[df_base["Date_label"] == sel1]

    with c2:
        opts2 = ["(All)"] + sorted([x for x in df1[c2_word].unique() if str(x).strip()])
        sel2 = st.selectbox(c2_word, opts2, index=0)

    df2 = df1 if sel2 == "(All)" else df1[df1[c2_word] == sel2]

    with c3:
        opts3 = ["(All)"] + sorted([x for x in df2[c3_word].unique() if str(x).strip()])
        sel3 = st.selectbox(c3_word, opts3, index=0)

    df3 = df2 if sel3 == "(All)" else df2[df2[c3_word] == sel3]

    with c4:
        opts4 = ["(All)"] + sorted([x for x in df3[c4_word].unique() if str(x).strip()])
        sel4 = st.selectbox(c4_word, opts4, index=0)

    df4 = df3 if sel4 == "(All)" else df3[df3[c4_word] == sel4]

    with c5:
        opts5 = ["(All)"] + sorted([x for x in df4[c5_word].unique() if str(x).strip()])
        sel5 = st.selectbox(c5_word, opts5, index=0)

    df5 = df4 if sel5 == "(All)" else df4[df4[c5_word] == sel5]

    with c6:
        text_search = st.text_input("Text search", value="", placeholder="Type to search Text...")

    df_view = df5
    if text_search.strip():
        s = text_search.strip().lower()
        df_view = df_view[df_view[search_word].astype(str).str.lower().str.contains(s, na=False)]

    st.caption(f"Rows: {len(df_view)}")


    df_grid = df_view.drop(['Date_dt','Date_label'],axis=1,errors='ignore')

    gb = GridOptionsBuilder.from_dataframe(df_grid)
    gb.configure_column("Date", flex=1, minWidth=70)
    gb.configure_column("Item", flex=1, minWidth=70)
    gb.configure_column("Source", flex=2, minWidth=100, wrapText=True)
    gb.configure_column("Chapter", flex=1, minWidth=70)
    gb.configure_column("Verse(s)", flex=1, minWidth=70)
    gb.configure_column("Text", flex=6, minWidth=300, wrapText=True, autoHeight=True)

    AgGrid(df_grid, gridOptions=gb.build(), height=800, fit_columns_on_grid_load=True)  
# -----------------------------------
# D Definitions
# -----------------------------------

elif page == "D Definitions":
    st.title("D Definitions")

    import pandas as pd
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode

    df_base = data_dict["google_definition_df"].copy()

    # Only convert actual NaN/None to ""
    df_base = df_base.fillna("")

    required = ["Process", "Categorization", "Word", "Definition"]
    missing = [c for c in required if c not in df_base.columns]
    if missing:
        st.error(f"google_definition_df is missing required columns: {missing}")
        st.stop()

    # ----------------------------
    # 1) Slicers
    # ----------------------------
    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        opts1 = ["(All)"] + sorted([x for x in df_base["Process"].astype(str).unique() if str(x).strip()])
        sel1 = st.selectbox("Process", opts1, index=0)

    df1 = df_base if sel1 == "(All)" else df_base[df_base["Process"].astype(str) == str(sel1)]

    with c2:
        opts2 = ["(All)"] + sorted([x for x in df1["Categorization"].astype(str).unique() if str(x).strip()])
        sel2 = st.selectbox("Categorization", opts2, index=0)

    df2 = df1 if sel2 == "(All)" else df1[df1["Categorization"].astype(str) == str(sel2)]

    with c3:
        opts3 = ["(All)"] + sorted([x for x in df2["Word"].astype(str).unique() if str(x).strip()])
        sel3 = st.selectbox("Word", opts3, index=0)

    df_view_full = df2 if sel3 == "(All)" else df2[df2["Word"].astype(str) == str(sel3)]
    st.caption(f"Rows: {len(df_view_full)}")

    # ----------------------------
    # 2) Grid (4 visible cols) + hidden _row_id
    # ----------------------------
    df_view_full = df_view_full.copy().reset_index(drop=False).rename(columns={"index": "_row_id"})

    visible_cols = ["Process", "Categorization", "Word", "Definition"]
    grid_df = df_view_full[["_row_id"] + visible_cols].copy()

    gb = GridOptionsBuilder.from_dataframe(grid_df)
    gb.configure_default_column(resizable=True, sortable=True, filter=True, wrapText=True, autoHeight=True)
    gb.configure_selection("single", use_checkbox=False)
    gb.configure_column("_row_id", hide=True)

    gb.configure_column("Process", width=120)
    gb.configure_column("Categorization", width=160)
    gb.configure_column("Word", width=160)
    gb.configure_column("Definition", width=520)

    grid_resp = AgGrid(
        grid_df,
        gridOptions=gb.build(),
        height=500,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
    )

    selected_rows = grid_resp.get("selected_rows", [])
    if selected_rows is None:
        selected_rows = []
    elif isinstance(selected_rows, pd.DataFrame):
        selected_rows = selected_rows.to_dict("records")

    # ----------------------------
    # 3) Details (HTML-ish rendering)
    # ----------------------------
    st.subheader("Details")

    if len(selected_rows) == 0:
        st.info("Select a row above to view full details.")
    else:
        row_id = selected_rows[0].get("_row_id", None)

        if row_id is None:
            st.warning("Selection did not return _row_id (unexpected).")
        else:
            full_row = df_view_full[df_view_full["_row_id"] == row_id].head(1)

            if full_row.empty:
                st.warning("Could not locate the full record for the selected row.")
            else:
                rec = full_row.iloc[0].fillna("")

                # Show Image (if present)
                if "Image" in rec.index:
                    img_url = str(rec["Image"]).strip()
                    if img_url:
                        st.image(img_url, caption="Image", width=320)

                # Render each field/value like your reference function
                # (Exclude helper + Image since already shown)
                exclude_fields = {"_row_id", "Image"}

                for field, value in rec.items():
                    if field in exclude_fields:
                        continue

                    v = "" if value is None else str(value).strip()

                    # Always show field, even if blank
                    if field.lower() == "link":
                        if v:
                            st.markdown(f"**{field}:** [Open Link]({v})")
                        else:
                            st.markdown(f"**{field}:**")
                    elif field.lower() in {"markdown", "latex"}:
                        st.markdown(f"**{field}:**")
                        if v:
                            try:
                                st.latex(v)
                            except Exception:
                                st.write(v)
                        else:
                            st.write("")
                    else:
                        st.markdown(f"**{field}:**")
                        st.write(v)

elif page == "Technical Notes":
    st.title("Technical Notes")

    df_base = data_dict['technical_notes'].copy()

    c1_word = "Program"
    c2_word = "Description"
    c3_word = "Classification"

    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        opts1 = ["(All)"] + sorted([x for x in df_base[c1_word].unique() if x.strip()])
        sel1 = st.selectbox(c1_word, opts1, index=0)

    df1 = df_base if sel1 == "(All)" else df_base[df_base[c1_word] == sel1]

    with c2:
        opts2 = ["(All)"] + sorted([x for x in df1[c2_word].unique() if x.strip()])
        sel2 = st.selectbox(c2_word, opts2, index=0)

    df2 = df1 if sel2 == "(All)" else df1[df1[c2_word] == sel2]

    with c3:
        opts3 = ["(All)"] + sorted([x for x in df2[c3_word].unique() if x.strip()])
        sel3 = st.selectbox(c3_word, opts3, index=0)

    df_view = df2 if sel3 == "(All)" else df2[df2[c3_word] == sel3]

    st.caption(f"Rows: {len(df_view)}")

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_column('Program', width=50, wrapText=True,autoHeight=True)
    gb.configure_column('Classification', width=50, wrapText=True,autoHeight=True)
    gb.configure_column('Command_Code', width=80, wrapText=True,autoHeight=True)
    gb.configure_column('Comments', width=80, wrapText=True,autoHeight=True)

    
    gb.configure_column(
        'Description',
        width=200,
        wrapText=True,
        autoHeight=True,  # makes row height expand to fit wrapped text
        cellStyle={
            "whiteSpace": "normal",
            "lineHeight": "1.2",
        },
    )
    gb.configure_default_column(resizable=True, sortable=True, filter=True)
    gb.configure_grid_options(domLayout="normal")
    grid_options = gb.build()

    AgGrid(
        df_view,
        gridOptions=grid_options,
        fit_columns_on_grid_load=True,
        height=800,
        theme="streamlit",
    )


elif page == "Frequency Summarization":
    st.title("Frequency Summarization")

    # ---- Source dropdown ----
    source_label_to_key = {
        "Consolidated File": "notes_def_summary",
        "Notes Page": "notes_summary",
        "Definition Page": "def_summary",
    }

    source_label = st.selectbox(
        "Source",
        list(source_label_to_key.keys()),
        index=0,
        key="def_source_select",
    )
    source_key = source_label_to_key[source_label]

    # ---- Dimension toggles ----
    show_process = st.checkbox("Include Process", value=True)
    show_category = st.checkbox("Include Categorization", value=True)
    show_word = st.checkbox("Include Word", value=False)

    # ---- Base data (immutable) ----
    df_base = data_dict[source_key].copy()

    # Safely drop Definition if present (won't error if missing)
    if "Definition" in df_base.columns:
        df_base = df_base.drop(columns=["Definition"])

    # Keep rows with a valid Word (Word is the atomic concept in your sheet)
    df_base = df_base[
        df_base["Word"].notna() &
        (df_base["Word"].astype(str).str.strip() != "")
    ].copy()

    # Column names
    COL_PROCESS = "Process"
    COL_CAT = "Categorization"
    COL_WORD = "Word"

    # ---- Filter UI (only for included dimensions) ----
    filter_cols = []
    if show_process:
        filter_cols.append(COL_PROCESS)
    if show_category:
        filter_cols.append(COL_CAT)
    if show_word:
        filter_cols.append(COL_WORD)

    filter_containers = st.columns(len(filter_cols)) if filter_cols else []

    df_filtered = df_base.copy()

    for col, container in zip(filter_cols, filter_containers):
        with container:
            opts = ["(All)"] + sorted(
                [x for x in df_filtered[col].dropna().astype(str).unique() if x.strip()]
            )
            sel = st.selectbox(
                col,
                opts,
                index=0,
                key=f"{source_key}_filter_{col}",  # unique per source
            )

        if sel != "(All)":
            df_filtered = df_filtered[df_filtered[col].astype(str) == sel]

    # ---- Build view dataframe ----
    df_view = df_filtered.copy()

    # ---- Determine deduplication grain ----
    dedupe_cols = []
    if show_process:
        dedupe_cols.append(COL_PROCESS)
    if show_category:
        dedupe_cols.append(COL_CAT)
    if show_word:
        dedupe_cols.append(COL_WORD)

    # Deduplicate on the visible grain (prevents dupes when rolling up)
    if dedupe_cols and len(dedupe_cols) < 3:
        df_view = df_view.drop_duplicates(subset=dedupe_cols).reset_index(drop=True)
    else:
        df_view = df_view.reset_index(drop=True)

    # ---- Keep only count columns relevant to current grain ----
    grain_to_counts = {
        (COL_PROCESS,): {"Process_Count"},
        (COL_CAT,): {"CAT_Count"},
        (COL_WORD,): {"Word_Count"},
        (COL_PROCESS, COL_CAT): {"Process_Count", "CAT_Count", "ProcessCAT_Count"},
        (COL_PROCESS, COL_WORD): {"Process_Count", "Word_Count"},
        (COL_CAT, COL_WORD): {"CAT_Count", "Word_Count"},
        (COL_PROCESS, COL_CAT, COL_WORD): {"Process_Count", "CAT_Count", "Word_Count", "ProcessCAT_Count"},
    }

    all_count_cols = {"Process_Count", "CAT_Count", "Word_Count", "ProcessCAT_Count"}
    grain_key = tuple(dedupe_cols)
    keep_counts = grain_to_counts.get(grain_key, set())

    drop_counts = [c for c in all_count_cols if (c in df_view.columns and c not in keep_counts)]
    if drop_counts:
        df_view = df_view.drop(columns=drop_counts)

    # ---- Drop hidden columns from the grid ----
    cols_to_drop = []
    if not show_process and COL_PROCESS in df_view.columns:
        cols_to_drop.append(COL_PROCESS)
    if not show_category and COL_CAT in df_view.columns:
        cols_to_drop.append(COL_CAT)
    if not show_word and COL_WORD in df_view.columns:
        cols_to_drop.append(COL_WORD)
        if "Word_Count" in df_view.columns:
            cols_to_drop.append("Word_Count")

    if cols_to_drop:
        df_view = df_view.drop(columns=cols_to_drop)

    st.caption(
        f"Source: {source_label} | Rows: {len(df_view)} | Grain: " +
        (" + ".join(dedupe_cols) if dedupe_cols else "(none)")
    )

    # ---- AgGrid ----
    gb = GridOptionsBuilder.from_dataframe(df_view)

    if show_process and COL_PROCESS in df_view.columns:
        gb.configure_column(COL_PROCESS, width=140, wrapText=True, autoHeight=True)
    if show_category and COL_CAT in df_view.columns:
        gb.configure_column(COL_CAT, width=160, wrapText=True, autoHeight=True)
    if show_word and COL_WORD in df_view.columns:
        gb.configure_column(COL_WORD, width=180, wrapText=True, autoHeight=True)

    # Count / metric columns (if present)
    for col, w in [
        ("Process_Count", 120),
        ("CAT_Count", 120),
        ("Word_Count", 120),
        ("ProcessCAT_Count", 160),
    ]:
        if col in df_view.columns:
            gb.configure_column(col, width=w, wrapText=True, autoHeight=True)

    gb.configure_default_column(resizable=True, sortable=True, filter=True)
    gb.configure_grid_options(domLayout="normal")
    grid_options = gb.build()

    AgGrid(
        df_view,
        gridOptions=grid_options,
        fit_columns_on_grid_load=True,
        height=800,
        theme="streamlit",
    )
