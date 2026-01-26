import pandas as pd
import numpy as np
import html
import textwrap
import random

import sys
sys.path.append("/Users/derekdewald/Documents/Python/Github_Repo/d_py_functions")

from list_processing import random_choice_from_uniform_list,random_uniform_normalized_list


def BinaryComplexEquivlancey(df, col, col1, new_column_name):
    try:
        # Try numeric comparison
        df[new_column_name] = np.where(
            (df[col].isna() & df[col1].isna()) |
            ((df[col].fillna(0) == 0) & df[col1].isna()) |
            ((df[col1].fillna(0) == 0) & df[col].isna()) |
            (df[col].fillna(0) == df[col1].fillna(0)),
            1, 0
        )
    except Exception:
        # Fallback to string comparison
        df[new_column_name] = np.where(
            (df[col].isna() & df[col1].isna()) |
            ((df[col].fillna('') == '') & df[col1].isna()) |
            ((df[col1].fillna('') == '') & df[col].isna()) |
            (df[col].fillna('').astype(str).str.strip().str.lower() ==
             df[col1].fillna('').astype(str).str.strip().str.lower()),
            1, 0
        )
    return df


def binary_column_creator(df,
                          column_name,
                          new_column_name=None,
                          value=0,
                          calculation='=',
                          balance_column=None):
    '''
    Function to Create a Binary Flag. 
    
    Updated to remove Dictionary capabilities. Which seems overtly complex and unncessary. 23Jul25
    
    '''
    
    if not new_column_name:
        flag = f"{column_name.upper()}_FLAG"
        bal = f"{column_name.upper()}_BAL"
    else:
        flag = f"{new_column_name.upper()}_FLAG"
        bal = f"{new_column_name.upper()}_BAL"
        
    if calculation=='>':
        df[flag] = np.where(df[column_name]>value,1,0)
    elif calculation =='=':
        df[flag] = np.where(df[column_name]==value,1,0)
    elif calculation =='<':
        df[flag] = np.where(df[column_name]<value,1,0)
    elif calculation =='isin':
        df[flag] = np.where(df[column_name].isin(value),1,0)
    elif calculation =='contains':
        df[flag] = np.where(df[column_name].str.contains(value),1,0)
    if balance_column:
        df[bal] = np.where(df[flag]==1,df[balance_column],0)

#from IPython.display import display, HTML

def notes_df_to_outline_html(
    df: pd.DataFrame,
    column_order=None):
    
    """

    Function to Take a Dataframe and convert it into A Structured Indented Point form Format. 
    Used for Clear Visualization of Notes.
    
    Parameters:
        df(df): Any DataFrame
        column_order(list): List of Columns to Include, in Order. If not defined, all will be included.
        print_(bool): Option as to whether you wish to directly Render a print out in the Python Session. Added because of Streamlit Error, need to suppress.

    Returns:
        str

    date_created:12-Dec-25
    date_last_modified: 18-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        from connections import d_google_sheet_to_csv
        df = import_d_google_sheet('Notes')
        notes_df_to_outline_html(df)

    Update: 
        Added display parameter to support Streamlit Adoption.

    """
    if column_order is None:
        column_order = df.columns.tolist()

    missing = [c for c in column_order if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    df1 = df[column_order].copy()

    def clean(x):
        if pd.isna(x):
            return ""
        return str(x).strip()

    last = [""] * len(column_order)

    html_ = """
    <style>
    .notes-container { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial; }
    .notes-item { line-height: 1.45; margin: 2px 0; }

    .notes-l0 { font-size: 18px; font-weight: 600; margin-left: 0px; }
    .notes-l1 { font-size: 16px; font-weight: 500; margin-left: 18px; }
    .notes-l2 { font-size: 14px; font-weight: 400; margin-left: 36px; }
    .notes-l3 { font-size: 13px; font-weight: 400; margin-left: 54px; opacity: 0.85; }
    .notes-l4 { font-size: 12px; font-weight: 400; margin-left: 72px; opacity: 0.8; }
    </style>

    <div class="notes-container">
    """

    for _, row in df1.iterrows():
        vals = [clean(row[c]) for c in column_order]
        if all(v == "" for v in vals):
            continue

        # Find first level where value changes
        change_level = None
        for i, v in enumerate(vals):
            if v and v != last[i]:
                change_level = i
                break

        # If nothing changes, show deepest non-blank value
        if change_level is None:
            for i in range(len(vals) - 1, -1, -1):
                if vals[i]:
                    change_level = i
                    break

        # Still nothing? (paranoia guard)
        if change_level is None:
            continue

        # Reset deeper levels when higher level changes (deeper only)
        for j in range(change_level + 1, len(last)):
            last[j] = ""

        # Render new values from change_level downward
        for i in range(change_level, len(vals)):
            v = vals[i]
            if not v:
                continue
            if v != last[i]:
                level = min(i, 4)  # cap style depth
                safe_v = html.escape(v) # Escape Function, not String.
                html_ += f'<div class="notes-item notes-l{level}">{safe_v}</div>\n'
                last[i] = v

    html_ += "</div>"
    html_ = textwrap.dedent(html_).lstrip()
    return html_

def final_dataset_for_markdown(notes=None,
                                  definitions=None,
                                  export_location='/Users/derekdewald/Documents/Python/Github_Repo/Streamlit/DataDictionary/'):
    
    '''
    
    Function which helps to combined Notes and Definitions into a Single Combined Representation which can ultimately be used as a Learning Reference Tools.

    How this function Works:
    It takes the two sheets and attemps to Consolidate them together to make Final Dataset, generated as d_learning_notes.csv.

    Approach is to Take Notes As the Framework and Distribute All Information into the Notes Sequentially in Specific Order so i can 
    Understand the structure and how to continue and utilize the Sheet.

    Principles: 
        - Define all types of Records in Notes and then Move them Out Definiton By Definition
        - Worry about Order at End.

    Step 1: Insert Records where Categorization = Definition directly into Sheet.
            With Modification:
                Where Word is also A Categorization, Update Categorization from Definition to Word, Change Word to Definition, 
                and Updated Definition to include WORD:

    Step 2: Take Everything else. Should be nothing remaining from Notes Category, so need to update Category, by moving all information 
            to the Right 1 column and consolidating Definition.
            
    Parameters:
        notes(df): DataFrame of D Notes as stored in: https://docs.google.com/spreadsheets/d/e/2PACX-1vSQF2lNc4WPeTRQ_VzWPkqSZp4RODFkbap8AqmolWp5bKoMaslP2oRVVG21x2POu_JcbF1tGRcBgodu/pub?output=csv
        definitions(df): DataFrame of D Definitions as stored in: https://docs.google.com/spreadsheets/d/e/2PACX-1vQq1-3cTas8DCWBa2NKYhVFXpl8kLaFDohg0zMfNTAU_Fiw6aIFLWfA5zRem4eSaGPa7UiQvkz05loW/pub?output=csv
        export_location(str): Location of where to Save CSV File. If blank, no CSV is made.

    date_created:20-Dec-25
    date_last_modified: 21-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        final_dataset_for_markdown()

    ##############

    Has been tested for a Single Value - Machine Learning. Need to Validate once extending.

    ##############

    
    '''

    from data_d_dicts import links

    try:
        len(notes)
    except:
        notes = pd.read_csv(links['google_notes_csv'])

    final_df = notes.copy()
    
    try:
        len(definitions)
    except:
        definitions = pd.read_csv(links['google_definition_csv'])

    temp_def = definitions[['Category','Categorization','Word','Definition']].copy()

    # Step 1
    # Create Definition only df and a Residual Definition DF res_def_df
    df_cat_is_definition = temp_def[temp_def['Categorization']=='Definition'].copy()

    
    
    cat_list = notes['Categorization'].unique().tolist()
    # When Word is also a Categorization, update so it is grouped correctly when sorting
    df_cat_is_definition['Categorization'] = np.where(df_cat_is_definition['Word'].isin(cat_list),df_cat_is_definition['Word'],df_cat_is_definition['Categorization'])
    
    mask = df_cat_is_definition["Categorization"] != "Definition"

    df_cat_is_definition.loc[mask, "Definition"] = (
        df_cat_is_definition.loc[mask, "Word"].astype(str)
        + ": "
        + df_cat_is_definition.loc[mask, "Definition"].astype(str)
    )

    df_cat_is_definition['Word'] = np.where(df_cat_is_definition['Categorization']!='Definition','Definition',df_cat_is_definition['Word'])
    res_def_df = temp_def[temp_def['Categorization']!='Definition'].copy()
    

    #Step 2
    
    def prepare_df_for_insert(df,notes=notes):
        df = df.copy()
        df['Definition'] = df['Word'] + ": " + df['Definition']
        df['Word'] = df["Categorization"].copy()
        df['Categorization'] = df["Category"].copy()
        df.drop('Category',axis=1,inplace=True)
        
        # Need to Identify Category. Should Primarily be from Categorization, might be from Word in Rare cases ("Regularization")
        df1 =  df[['Categorization']].drop_duplicates().merge(notes[['Category','Categorization']].drop_duplicates(),on='Categorization',how='left')    
        return df.merge(df1,on=['Categorization'],how='left')

    res_def_df = prepare_df_for_insert(res_def_df)
    
    # Insert 
    final_df = pd.concat([final_df,df_cat_is_definition,res_def_df])

    ################
    
    # Ranking for Sort ORder
    rank_df1 = notes.drop_duplicates('Category')[['Category']].reset_index(drop=True).reset_index().rename(columns={'index':"CY_RANK"})
    rank_df2 = notes.drop_duplicates(['Category','Categorization'])[['Category','Categorization']].reset_index(drop=True).reset_index().rename(columns={'index':"CZ_RANK"})
    rank_df2['CZ_RANK'] = rank_df2['CZ_RANK'] + 1
    rank_df3 = notes.drop_duplicates(['Category','Categorization','Word']).reset_index(drop=True).reset_index().drop('Definition',axis=1).rename(columns={'index':'WORD_RANK'})
    rank_df3['WORD_RANK'] = rank_df3['WORD_RANK'] + 1

    final_df = final_df.merge(rank_df1,on=['Category'],how='left').merge(rank_df2,on=['Category','Categorization'],how='left').merge(rank_df3,on=['Category','Categorization','Word'],how='left')
    final_df['CZ_RANK'] = np.where(final_df['Categorization']=='Definition',0,final_df['CZ_RANK'])
    final_df['WORD_RANK'] = np.where(final_df['Word']=='Definition',0,final_df['WORD_RANK'])
    
    # Sort Data Frame to Desired Order, then Remove Ranking Values
    final_df =  final_df.sort_values(['CY_RANK','CZ_RANK','WORD_RANK']).drop(['CY_RANK','CZ_RANK','WORD_RANK'],axis=1)

    # Remove Duplicate Values where Records from Definitions has merged in values and there more than 1 Word Records and one has a Blank Value.
    final_df['COUNT'] = final_df.groupby(['Category','Categorization','Word']).transform('size')
    final_df = final_df[~((final_df['COUNT']>1)&(final_df['Definition']==""))]

    if export_location:
        final_df.to_csv(f"{export_location}d_learning_notes.csv",index=False)
    
    return final_df


def random_uniform_normalized_df(unique_records,
                                 name='Example',
                                 skew=1.25,
                                 **kwargs):
    '''
    Create a Dataframe (which is a series of n * 1) of Random Values for purposes of creating a Random Distribution DataFrame.
    Kwargs can be used to create New Columns. Kwargs should be Lists of distribution Frequencies, to create new random Columns (Not cdf).

    Parameters:
        unique records(int): Number, representing the number of random columns to be included in the output DF.
        name(str): Name of Column to Included (values will be numbered).
        skew(float): If Data is to have a skewed distribution, 1 will be normal uniform (mean=1,std_dev=0).
        **kwargs: Should be List of values equalling 1, to create a new random value.

    Returns:
        Object Type

    date_created:29-Dec-25
    date_last_modified: 29-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        random_uniform_normalized_df(unique_records=40,name='BRANCHNAME',LEGACY=[.5,.15,.3,.05])
    
    '''
    obs_name_list = [f'{name} {x+1}' for x in range(0,unique_records)]
    dist_perc = random_uniform_normalized_list(unique_records,skew=skew)
    
    final_df = pd.DataFrame()
    
    for obs in range(0,unique_records):
        obs_name = obs_name_list[obs]
        perc_ = dist_perc[obs]
        temp_df = pd.DataFrame([[obs_name,perc_]],columns=[name,'PERC_'])
        final_df = pd.concat([final_df,temp_df])
            
    for kwarg_name, kwarg_value in kwargs.items():
        temp_df = random_choice_from_uniform_list(1000,name=kwarg_name,list_distribution=kwarg_value,return_value='df')
        final_df = final_df.reset_index(drop=True).merge(temp_df,left_index=True,right_index=True,how='left')
    
    return final_df

def df_to_dict(df,key,value):
    
    '''
    Function to Simply Convert A DF into a Dictionary.
    Takes 2 Arguments, and converts them into a DF of the format {key:value}

    Parameters:
        df (df): Any DataFrame
        key (str): String representing Column Name for Dictionary Key
        value(str): String representing Column Name for Value Key
        
    Returns:
        df

    date_created:12-Dec-25
    date_last_modified: 12-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        from data_d_strings import google_mapping_sheet_csv
        df = pd.read_csv(google_mapping_sheet_csv)
        df_to_dict(df,'Definition','CSV')

    '''

    temp_df = df[[key,value]].copy()

    return df[[key,value]].set_index(key).to_dict()[value]


def replicate_df_row(df,records=5):
    
    '''
    Function which Replicates a single row DataFrame for the purposes of Multiplying it against a larger row.
    Function written using tile, which is a C based language, and considerably faster than straight using nunpy vectorized Calculations.

    Parameters:
        df(dataframe): DataFrame which you wish to extend, should be a Single Row, but techincally it will duplicate any size
        records(int): Number of times you wish DF to be duplicated, ideally it should be len(other_df) to which you want to multiply

    Returns:
        df

    date_created:30-Dec-25
    date_last_modified: 30-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        df = pd.DataFrame([[1,2,3]],columns=['A','B','C'])
        replicate_row(df)
    
    '''
    
    row = df.to_numpy()
    columns = df.columns.tolist()

    # Repeat row N times using NumPy
    data = np.tile(row, (records, 1))  # shape (N, len(row))
    return pd.DataFrame(data, columns=columns)

def tranpose_df(df, index, columns=None):
    '''

    Transposes a non-time-series DataFrame from wide to long format by melting specified columns.

    This is especially useful for flattening columns into a single column to support tools 
    like Power BI, where long format enables dynamic pivoting and aggregation.


    Parameters:
        df (DataFrame): The input pandas DataFrame.
        index (list): Columns to retain as identifiers (will remain unchanged).
        columns (list): Columns to unpivot into key-value pairs.

    Returns:
        DataFrame: A long-format DataFrame with 'variable' and 'value' columns.

    date_created:1-Jan-24
    date_last_modified: 30-Dec-25
    classification:TBD
    sub_classification:TBD
    usage:
        Example Function Call

    '''
    if not columns:
        columns = [col for col in df.columns if col not in index]
    return df.melt(id_vars=index, value_vars=columns)   



def export_to_excel(df,
                    file_name='python_excel_file.xlsx',
                    sheet_name= "Sheet1",
                    default_max_width= 30,
                    long_columns=[],
                    long_max_width= 80):
    """
    Function Created to Export Data to Excel, with increased Control over the output, including adding a sheet name and attempting to 
    format the columns, which would be the primary use over simple .to_excel. 

    Given CSV format does not have explicit memory, there is no benefit when requirement dictate a CSV file.


    Parameters:
    sheet_name(str): Name of Sheet to be utilized in Excel.
    default_max_width(int): Default Column Width
    long_columns (list): Any column which would be requested to have a default column width beyond 30, include in thelist
    long_max_width(int): Max width for long_columns
    
    
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from typing import Iterable, Optional
    from io import BytesIO
    
    min_width = 10
    padding = 2
    wrap_vertical_align= "top"
    freeze_header= True,

    long_columns_set = set(long_columns or [])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        safe_sheet = sheet_name[:31]  # Excel sheet name limit
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.sheets[safe_sheet]

        if freeze_header:
            ws.freeze_panes = "A2"

        # Predefine alignments (reuse objects)
        wrap_align = Alignment(wrap_text=True, vertical=wrap_vertical_align)
        no_wrap_align = Alignment(wrap_text=False, vertical=wrap_vertical_align)

        for i, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)

            # Compute max string length in this column (including header)
            ser = df[col].astype(str).fillna("")
            max_len = max(len(str(col)), int(ser.map(len).max()) if len(ser) else 0)

            # Choose cap
            cap = long_max_width if col in long_columns_set else default_max_width

            # Proposed width (with padding)
            proposed = max_len + padding

            # Final width with min + cap
            final_width = max(min_width, min(proposed, cap))
            ws.column_dimensions[col_letter].width = final_width

            # Wrap if we had to cap (meaning content would exceed the allowed width)
            should_wrap = proposed > cap
            if should_wrap:
                # Apply wrap to entire column, incl header
                for cell in ws[col_letter]:
                    cell.alignment = wrap_align
            else:
                # Optional: set vertical alignment consistently
                for cell in ws[col_letter]:
                    cell.alignment = no_wrap_align

    excel_bytes = buffer.getvalue()
    
    with open(file_name, "wb") as f:
        f.write(excel_bytes)


def transpose_df(df, index, columns=None):
    '''
    Transposes a non-time-series DataFrame from wide to long format by melting specified columns.

    This is especially useful for flattening columns into a single column to support tools 
    like Power BI, where long format enables dynamic pivoting and aggregation.

    Parameters:
        df (DataFrame): The input pandas DataFrame.
        index (list): Columns to retain as identifiers (will remain unchanged).
        columns (list): Columns to unpivot into key-value pairs.

    Returns:
        DataFrame: A long-format DataFrame with 'variable' and 'value' columns.

        
        Definition of Function

    date_created: 1-JUL-25
    date_last_modified: 1-JUL-25
    classification:TBD
    sub_classification:TBD
    usage:
        
        

    '''
    if not columns:
        columns = [col for col in df.columns if col not in index]
    return df.melt(id_vars=index, value_vars=columns)   